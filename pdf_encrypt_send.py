#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF加密并发送邮件脚本 (Excel配置版)
功能：将指定文件夹内的PDF文件加密后，通过163邮箱发送
配置：使用Excel表格管理文件夹、密码、邮箱对应关系
"""

import json
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path
import tempfile
import shutil
from colorama import init, Fore, Style
from openpyxl import load_workbook

# 初始化colorama（Windows支持）
init(autoreset=True)

def print_success(msg):
    """打印成功消息"""
    print(f"{Fore.GREEN}✓ {msg}{Style.RESET_ALL}")

def print_error(msg):
    """打印错误消息"""
    print(f"{Fore.RED}✗ {msg}{Style.RESET_ALL}")

def print_info(msg):
    """打印信息"""
    print(f"{Fore.CYAN}ℹ {msg}{Style.RESET_ALL}")

def print_warning(msg):
    """打印警告"""
    print(f"{Fore.YELLOW}⚠ {msg}{Style.RESET_ALL}")

def load_config():
    """加载JSON配置文件"""
    try:
        config_path = Path(__file__).parent / 'config.json'
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        print_success(f"配置文件加载成功: {config_path}")
        return config
    except FileNotFoundError:
        print_error("配置文件 config.json 不存在！")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print_error(f"配置文件格式错误: {e}")
        sys.exit(1)
    except Exception as e:
        print_error(f"加载配置文件失败: {e}")
        sys.exit(1)

def load_excel_config(excel_path):
    """
    从Excel文件加载配置

    Args:
        excel_path: Excel文件路径

    Returns:
        dict: 文件夹配置字典
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        folders_config = {}

        # 从第2行开始读取（第1行是表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过空行
            if not row[0]:
                continue

            folder_name = str(row[0]).strip()
            password = str(row[1]).strip() if row[1] else ""
            recipient_email = str(row[2]).strip() if row[2] else ""
            recipient_name = str(row[3]).strip() if row[3] else folder_name

            # 验证必填字段
            if not folder_name or not password or not recipient_email:
                print_warning(f"跳过不完整的行: {row}")
                continue

            folders_config[folder_name] = {
                "password": password,
                "recipient_email": recipient_email,
                "recipient_name": recipient_name
            }

        print_success(f"Excel配置加载成功: {excel_path}")
        print_info(f"共加载 {len(folders_config)} 个文件夹配置")

        return folders_config

    except FileNotFoundError:
        print_error(f"Excel配置文件不存在: {excel_path}")
        sys.exit(1)
    except Exception as e:
        print_error(f"读取Excel配置失败: {e}")
        sys.exit(1)

def encrypt_pdf(input_path, output_path, password):
    """
    加密PDF文件

    Args:
        input_path: 输入PDF路径
        output_path: 输出PDF路径
        password: 加密密码

    Returns:
        bool: 是否成功
    """
    try:
        import pikepdf

        with pikepdf.open(input_path) as pdf:
            pdf.save(output_path, encryption=pikepdf.Encryption(
                user=password,
                owner=password,
                R=6  # 使用AES-256加密
            ))
        return True
    except Exception as e:
        print_error(f"加密失败 [{input_path.name}]: {e}")
        return False

def get_pdf_files(folder_path):
    """获取文件夹中的所有PDF文件"""
    pdf_files = list(folder_path.glob('*.pdf'))
    # 递归查找子文件夹中的PDF
    pdf_files.extend(folder_path.glob('**/*.pdf'))
    # 去重
    pdf_files = list(set(pdf_files))
    return pdf_files

def send_email(smtp_config, recipient_email, recipient_name, subject, body, attachments):
    """
    发送邮件

    Args:
        smtp_config: SMTP配置
        recipient_email: 收件人邮箱
        recipient_name: 收件人姓名
        subject: 邮件主题
        body: 邮件正文
        attachments: 附件列表（Path对象）

    Returns:
        bool: 是否成功
    """
    try:
        # 创建邮件
        msg = MIMEMultipart()
        # 163邮箱要求From字段必须与登录邮箱完全一致
        msg['From'] = smtp_config['sender_email']
        msg['To'] = recipient_email
        msg['Subject'] = subject

        # 添加正文
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # 添加附件
        for attachment_path in attachments:
            with open(attachment_path, 'rb') as f:
                part = MIMEApplication(f.read(), Name=attachment_path.name)
                part['Content-Disposition'] = f'attachment; filename="{attachment_path.name}"'
                msg.attach(part)

        # 连接SMTP服务器并发送
        print_info(f"正在连接163邮箱服务器...")
        with smtplib.SMTP_SSL(smtp_config['server'], smtp_config['port']) as server:
            server.login(smtp_config['sender_email'], smtp_config['sender_password'])
            server.send_message(msg)

        return True
    except smtplib.SMTPAuthenticationError:
        print_error("163邮箱认证失败！请检查邮箱账号和授权码是否正确")
        print_warning("提示：163邮箱需要使用授权码，不是邮箱密码！")
        return False
    except Exception as e:
        print_error(f"发送邮件失败: {e}")
        return False

def process_folder(folder_name, folder_config, smtp_config, source_path):
    """
    处理单个文件夹

    Args:
        folder_name: 文件夹名称
        folder_config: 文件夹配置
        smtp_config: SMTP配置
        source_path: PDF源文件夹路径

    Returns:
        bool: 是否成功
    """
    print(f"\n{Fore.YELLOW}{'='*60}{Style.RESET_ALL}")
    print_info(f"开始处理: {folder_name}")
    print(f"{Fore.YELLOW}{'='*60}{Style.RESET_ALL}")

    # 检查文件夹是否存在
    folder_path = Path(source_path) / folder_name
    if not folder_path.exists():
        print_error(f"文件夹不存在: {folder_path}")
        return False

    # 获取PDF文件
    pdf_files = get_pdf_files(folder_path)
    if not pdf_files:
        print_warning(f"文件夹中没有PDF文件: {folder_path}")
        return False

    print_info(f"找到 {len(pdf_files)} 个PDF文件")

    # 创建临时文件夹存放加密后的PDF
    temp_dir = Path(tempfile.mkdtemp())
    encrypted_files = []
    failed_files = []

    try:
        # 加密每个PDF
        password = folder_config['password']
        print_info(f"使用密码加密: {password}")

        for i, pdf_file in enumerate(pdf_files, 1):
            print_info(f"[{i}/{len(pdf_files)}] 正在加密: {pdf_file.name}")

            # 生成加密后的文件名
            encrypted_filename = f"加密_{pdf_file.name}"
            encrypted_path = temp_dir / encrypted_filename

            # 加密
            if encrypt_pdf(pdf_file, encrypted_path, password):
                encrypted_files.append(encrypted_path)
                print_success(f"加密成功: {pdf_file.name}")
            else:
                failed_files.append(pdf_file.name)

        # 如果有加密失败的文件
        if failed_files:
            print_warning(f"以下文件加密失败: {', '.join(failed_files)}")

        # 如果没有成功加密的文件
        if not encrypted_files:
            print_error("没有成功加密的文件，跳过发送邮件")
            return False

        # 发送邮件
        print_info(f"准备发送邮件到: {folder_config['recipient_email']}")

        subject = f"{folder_name} - 加密报告文件"
        body = f"""您好，{folder_config['recipient_name']}：

这是本月的报告文件，共 {len(encrypted_files)} 个PDF文件。

所有文件已加密，密码为: {password}

请及时查收！

---
此邮件由系统自动发送，请勿回复。
"""

        if send_email(smtp_config, folder_config['recipient_email'],
                     folder_config['recipient_name'], subject, body, encrypted_files):
            print_success(f"邮件发送成功！({len(encrypted_files)} 个附件)")
            return True
        else:
            return False

    finally:
        # 清理临时文件夹
        try:
            shutil.rmtree(temp_dir)
            print_info("临时文件已清理")
        except Exception as e:
            print_warning(f"清理临时文件失败: {e}")

def main():
    """主函数"""
    print(f"\n{Fore.CYAN}{'='*60}")
    print(f"{' '*15}PDF加密邮件发送系统")
    print(f"{' '*18}Excel配置版")
    print(f"{'='*60}{Style.RESET_ALL}\n")

    # 加载JSON配置
    config = load_config()

    # 验证邮箱配置
    if config['smtp']['sender_email'] == "your_email@163.com":
        print_error("请先修改 config.json 中的163邮箱配置！")
        sys.exit(1)

    # 加载Excel配置
    script_dir = Path(__file__).parent
    excel_path = script_dir / config['excel_config_file']

    if not excel_path.exists():
        print_error(f"Excel配置文件不存在: {excel_path}")
        print_info("请确保 配置表.xlsx 文件在脚本同一目录下")
        sys.exit(1)

    folders_config = load_excel_config(excel_path)

    if not folders_config:
        print_error("Excel配置表中没有有效的配置数据！")
        sys.exit(1)

    # 检查源文件夹
    source_path = Path(config['pdf_source_path'])
    if not source_path.exists():
        print_error(f"PDF源文件夹不存在: {source_path}")
        sys.exit(1)

    print_success(f"PDF源文件夹: {source_path}")

    # 处理每个文件夹
    total = len(folders_config)
    success_count = 0
    failed_folders = []

    for i, (folder_name, folder_config) in enumerate(folders_config.items(), 1):
        print(f"\n{Fore.MAGENTA}>>> 进度: {i}/{total} <<<{Style.RESET_ALL}")

        if process_folder(folder_name, folder_config, config['smtp'], source_path):
            success_count += 1
        else:
            failed_folders.append(folder_name)

    # 总结
    print(f"\n{Fore.CYAN}{'='*60}")
    print(f"{' '*20}处理完成")
    print(f"{'='*60}{Style.RESET_ALL}\n")

    print_info(f"总共: {total} 个文件夹")
    print_success(f"成功: {success_count} 个")

    if failed_folders:
        print_error(f"失败: {len(failed_folders)} 个")
        print_warning(f"失败的文件夹: {', '.join(failed_folders)}")

    print(f"\n{Fore.CYAN}按任意键退出...{Style.RESET_ALL}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print_warning("\n\n用户中断操作")
        sys.exit(0)
    except Exception as e:
        print_error(f"\n程序异常: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
